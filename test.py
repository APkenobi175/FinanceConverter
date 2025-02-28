import pandas as pd
import os
import sys
from datetime import datetime
from openpyxl import load_workbook


def categorize_expenses(converted_file, categorized_file):
    # Define valid categories and their corresponding spreadsheet sections
    category_ranges = {
        "Rent": "D3:F22",
        "Utilities": "G3:I22",
        "Car Expenses": "J3:L22",
        "Subscriptions": "M3:O22",
        "Taxes": "P3:R22",
        "Income": "S3:U22",
        "Savings": "A25:D50",
        "School": "D25:F50",
        "Groceries": "G25:I50",
        "Other": "J25:L50",
        "Fun": "M25:O50",
        "Eating Out": "P25:R100",
        "Retirement": "S25:U50"
    }

    # Load the converted expenses spreadsheet
    df = pd.read_excel(converted_file)

    if os.path.exists(categorized_file):
        wb = load_workbook(categorized_file)
    else:
        print("Error: Categorized spreadsheet does not exist.")
        sys.exit(1)

    while not df.empty:
        print("\nRemaining Expenses:")
        print(df.head())

        # Get the first expense to categorize
        row = df.iloc[0]
        expense_name = row["Expense"]
        amount = row["Amount"]
        date = row["Date"]

        # Ask if the user wants to disregard the expense
        disregard = input(f"Do you want to disregard '{expense_name} (${amount})'? (y/n, Enter 'q' to quit): ")
        if disregard.lower() == 'y':
            df = df.iloc[1:].reset_index(drop=True)
            df.to_excel(converted_file, index=False)  # Save remaining expenses
            continue
        elif disregard.lower() == 'q':
            print("Quitting and saving progress...")
            break

        # Ask for category
        while True:
            category = input(f"Enter category for '{expense_name}' (Enter 'q' to quit): ")
            if category.lower() == 'q':
                print("Quitting and saving progress...")
                return
            if category in category_ranges:
                break
            print("Error: Invalid category, enter a valid category.")

        # Ask for nickname
        nickname = input(f"Enter a nickname for '{expense_name}' (Enter 'q' to quit): ")
        if nickname.lower() == 'q':
            print("Quitting and saving progress...")
            return

        # Determine the month-year sheet name
        try:
            date_obj = datetime.strptime(str(date), "%m/%d/%Y")  # Adjust based on actual date format
            sheet_name = date_obj.strftime("%B %Y")
        except ValueError:
            print(f"Skipping invalid date format: {date}")
            continue

        # Ensure the sheet exists
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        # Find the next available row in the designated category section
        start_col, start_row = category_ranges[category].split(":")[0][0], int(
            category_ranges[category].split(":")[0][1:])
        for row_idx in range(start_row, 100):  # Arbitrary upper bound to find the next empty row
            if ws[f"{start_col}{row_idx}"].value is None:
                ws[f"{start_col}{row_idx}"] = nickname
                ws[f"{chr(ord(start_col) + 1)}{row_idx}"] = amount
                ws[f"{chr(ord(start_col) + 2)}{row_idx}"] = date
                break

        # Remove categorized row
        df = df.iloc[1:].reset_index(drop=True)

        # Save progress
        wb.save(categorized_file)
        df.to_excel(converted_file, index=False)  # Save remaining expenses

    print("\nCategorization complete. All changes have been saved.")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python script.py <converted_expenses.xlsx> <categorized_expenses.xlsx>")
        sys.exit(1)

    converted_expenses_file = sys.argv[1]
    categorized_expenses_file = sys.argv[2]
    categorize_expenses(converted_expenses_file, categorized_expenses_file)
